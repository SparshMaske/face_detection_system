import os
import csv
import shutil
from datetime import datetime
from glob import glob
from typing import List, Optional, Tuple

from flask import current_app
from models.visitor import Visitor, VisitorSession
from models.staff import Staff
from sqlalchemy import or_
import cv2

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    REPORTLAB_AVAILABLE = True
except ModuleNotFoundError:
    REPORTLAB_AVAILABLE = False

try:
    from fpdf import FPDF
except ModuleNotFoundError:
    FPDF = None

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ModuleNotFoundError:
    OPENPYXL_AVAILABLE = False

class ReportGenerator:
    def __init__(self):
        self.reports_dir = current_app.config['REPORTS_FOLDER']
        os.makedirs(self.reports_dir, exist_ok=True)
        self.visitor_reports_dir = os.path.join(self.reports_dir, 'visitors')
        os.makedirs(self.visitor_reports_dir, exist_ok=True)
        # Keep multiple cascades for better fallback detection coverage.
        self._face_cascades = []
        for cascade_name in ('haarcascade_frontalface_default.xml', 'haarcascade_frontalface_alt2.xml'):
            try:
                cascade_path = os.path.join(cv2.data.haarcascades, cascade_name)
                cascade = cv2.CascadeClassifier(cascade_path)
                if not cascade.empty():
                    self._face_cascades.append(cascade)
            except Exception:
                continue

    @staticmethod
    def _safe_text(value):
        if value is None:
            return ''
        text = str(value)
        return text.encode('latin-1', 'replace').decode('latin-1')

    @staticmethod
    def _ensure_pdf_backend():
        if REPORTLAB_AVAILABLE:
            return
        if FPDF is None:
            raise ModuleNotFoundError('No PDF backend installed (reportlab/fpdf)')

    @staticmethod
    def _ensure_excel_backend():
        if not OPENPYXL_AVAILABLE:
            raise ModuleNotFoundError('openpyxl')

    @staticmethod
    def _parse_datetime(value, is_end=False):
        if not value:
            raise ValueError('Date is required')
        parsed = datetime.fromisoformat(value)
        # If only a date is provided for end bound, include full day.
        if is_end and len(value) <= 10:
            parsed = parsed.replace(hour=23, minute=59, second=59, microsecond=999999)
        return parsed

    @staticmethod
    def _format_duration(seconds):
        total = max(0, int(seconds or 0))
        hours, remainder = divmod(total, 3600)
        minutes, secs = divmod(remainder, 60)
        return f"{hours:02d}:{minutes:02d}:{secs:02d}"

    @staticmethod
    def _sanitize_token(raw_value):
        text = str(raw_value or '').strip()
        if not text:
            return 'event'
        cleaned = []
        for ch in text:
            if ch.isalnum() or ch in ('-', '_'):
                cleaned.append(ch)
            elif ch.isspace():
                cleaned.append('_')
            else:
                cleaned.append('_')
        normalized = ''.join(cleaned).strip('_')
        return normalized or 'event'

    def _event_artifacts_dir(self, event_name):
        safe_event_name = self._sanitize_token(event_name)
        events_root = os.path.join(self.reports_dir, 'events')
        os.makedirs(events_root, exist_ok=True)
        event_dir = os.path.join(events_root, safe_event_name)
        os.makedirs(event_dir, exist_ok=True)
        return event_dir

    def _collect_management_payload(self, event_end: datetime) -> List[dict]:
        query = Staff.query.order_by(Staff.name.asc())
        if event_end is not None:
            query = query.filter(Staff.created_at <= event_end)
        rows = []
        for staff in query.all():
            rows.append({
                'staff_id': staff.staff_id or '',
                'name': staff.name or '',
                'department': staff.department or '-',
                'position': staff.position or '-',
                'status': 'Active' if staff.is_active else 'Inactive',
                'created_at': staff.created_at.strftime('%Y-%m-%d %H:%M:%S') if staff.created_at else '',
            })
        return rows

    def _save_event_data_csv(
        self,
        event_dir,
        event_name,
        event_id,
        start_date,
        end_date,
        visitors_payload,
        management_payload,
    ):
        safe_event_name = self._sanitize_token(event_name)
        suffix = self._sanitize_token(event_id) if event_id else datetime.now().strftime('%Y%m%d_%H%M%S')
        csv_filename = f"{safe_event_name}_{suffix}_data.csv"
        csv_path = os.path.join(event_dir, csv_filename)

        with open(csv_path, 'w', encoding='utf-8', newline='') as fp:
            writer = csv.writer(fp)
            writer.writerow([
                'record_type',
                'event_name',
                'event_id',
                'period_start',
                'period_end',
                'id',
                'name',
                'department',
                'position',
                'status',
                'date',
                'first_in_time',
                'last_out_time',
                'duration',
            ])

            for visitor_item in visitors_payload:
                writer.writerow([
                    'visitor',
                    event_name or '',
                    event_id or '',
                    start_date,
                    end_date,
                    visitor_item.get('visitor_id', ''),
                    '',
                    '',
                    '',
                    '',
                    visitor_item.get('date', ''),
                    visitor_item.get('first_in', ''),
                    visitor_item.get('last_out', ''),
                    visitor_item.get('duration', ''),
                ])

            for staff_item in management_payload:
                writer.writerow([
                    'management',
                    event_name or '',
                    event_id or '',
                    start_date,
                    end_date,
                    staff_item.get('staff_id', ''),
                    staff_item.get('name', ''),
                    staff_item.get('department', ''),
                    staff_item.get('position', ''),
                    staff_item.get('status', ''),
                    '',
                    '',
                    '',
                    '',
                ])

        return csv_path

    def _resolve_visitor_image_candidates(self, visitor, prefer_first=False, reference_time=None):
        upload_root = current_app.config.get('UPLOAD_FOLDER')
        visitor_root = current_app.config.get('VISITOR_UPLOAD_FOLDER')

        candidates = []
        images = list(visitor.images or [])
        if prefer_first:
            images.sort(key=lambda item: item.captured_at or datetime.max)
            if reference_time:
                # Pick image nearest to first appearance while preferring earlier captures.
                def _rank(item):
                    captured = item.captured_at or datetime.max
                    if captured <= reference_time:
                        return (0, (reference_time - captured).total_seconds())
                    return (1, (captured - reference_time).total_seconds())

                images = sorted(images, key=_rank)
        else:
            # Default behavior for other flows: newest image first.
            images.sort(key=lambda item: item.captured_at or datetime.min, reverse=True)

        for item in images:
            if item.image_path:
                candidates.append(item.image_path)

        if visitor.primary_image_path:
            candidates.append(visitor.primary_image_path)

        resolved = []
        seen = set()
        for raw_path in candidates:
            if not raw_path:
                continue

            # Support absolute file path records directly.
            if os.path.isabs(raw_path) and os.path.exists(raw_path):
                normalized_abs = os.path.abspath(raw_path)
                if normalized_abs not in seen:
                    seen.add(normalized_abs)
                    resolved.append(normalized_abs)
                continue

            normalized = raw_path.replace('\\', '/')
            if normalized.startswith('/'):
                normalized = normalized.lstrip('/')
            if normalized.startswith('static/'):
                normalized = normalized[len('static/'):]
            if normalized.startswith('uploads/'):
                normalized = normalized[len('uploads/'):]

            possible_paths = []
            if upload_root:
                possible_paths.append(os.path.join(upload_root, normalized))
            if visitor_root:
                possible_paths.append(os.path.join(visitor_root, os.path.basename(normalized)))
            possible_paths.append(os.path.join(self.visitor_reports_dir, 'snapshots', os.path.basename(normalized)))

            for abs_path in possible_paths:
                if abs_path and os.path.exists(abs_path):
                    normalized_abs = os.path.abspath(abs_path)
                    if normalized_abs not in seen:
                        seen.add(normalized_abs)
                        resolved.append(normalized_abs)
                    break
        return resolved

    def _resolve_visitor_image_path(self, visitor, prefer_first=False, reference_time=None):
        candidates = self._resolve_visitor_image_candidates(
            visitor,
            prefer_first=prefer_first,
            reference_time=reference_time,
        )
        return candidates[0] if candidates else None

    def _detect_primary_face_bbox(self, image) -> Optional[Tuple[int, int, int, int]]:
        if image is None:
            return None
        try:
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        except Exception:
            return None

        # Try stronger and then more permissive detector settings.
        params = [
            {'scaleFactor': 1.10, 'minNeighbors': 5, 'minSize': (40, 40)},
            {'scaleFactor': 1.06, 'minNeighbors': 4, 'minSize': (32, 32)},
        ]
        best = None
        best_area = 0
        for cascade in self._face_cascades:
            for cfg in params:
                try:
                    faces = cascade.detectMultiScale(gray, **cfg)
                except Exception:
                    faces = []
                for (x, y, w, h) in faces:
                    area = int(w) * int(h)
                    if area > best_area:
                        best_area = area
                        best = (int(x), int(y), int(w), int(h))
                if best is not None:
                    return best
        return best

    def _detect_primary_face_bbox_with_retries(self, image) -> Optional[Tuple[int, int, int, int]]:
        """
        More robust face lookup for report snapshots:
        1) original image
        2) equalized grayscale image
        3) upscaled image (helps small/soft faces)
        """
        bbox = self._detect_primary_face_bbox(image)
        if bbox is not None:
            return bbox

        try:
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            eq = cv2.equalizeHist(gray)
            eq_bgr = cv2.cvtColor(eq, cv2.COLOR_GRAY2BGR)
            bbox = self._detect_primary_face_bbox(eq_bgr)
            if bbox is not None:
                return bbox
        except Exception:
            pass

        try:
            h, w = image.shape[:2]
            if min(h, w) < 900:
                scale = 1.55
                up = cv2.resize(
                    image,
                    (max(1, int(w * scale)), max(1, int(h * scale))),
                    interpolation=cv2.INTER_CUBIC,
                )
                up_bbox = self._detect_primary_face_bbox(up)
                if up_bbox is not None:
                    x, y, fw, fh = up_bbox
                    return (
                        int(x / scale),
                        int(y / scale),
                        int(fw / scale),
                        int(fh / scale),
                    )
        except Exception:
            pass
        return None

    def _is_head_shoulder_crop(self, crop) -> bool:
        """
        Validate that the crop looks like a portrait (head to shoulder),
        not a torso-only or background-only patch.
        """
        if crop is None or getattr(crop, 'size', 0) == 0:
            return False
        face_bbox = self._detect_primary_face_bbox_with_retries(crop)
        if face_bbox is None:
            return False

        x, y, w, h = face_bbox
        ch, cw = crop.shape[:2]
        if cw <= 0 or ch <= 0 or w <= 0 or h <= 0:
            return False

        face_ratio = float(w * h) / float(cw * ch)
        # Head should be clearly visible but not occupy the whole crop.
        if face_ratio < 0.08 or face_ratio > 0.62:
            return False

        # Face center should sit in upper/middle portion for head-shoulder framing.
        cy = float(y + (h / 2.0)) / float(ch)
        if cy < 0.16 or cy > 0.58:
            return False
        return True

    def _prepare_face_to_shoulder_snapshot(self, image_path, visitor_code, allow_center_fallback=False):
        if not image_path or not os.path.exists(image_path):
            return None
        try:
            img = cv2.imread(image_path)
            if img is None:
                return None
            face_bbox = self._detect_primary_face_bbox_with_retries(img)
            if face_bbox is None:
                if not allow_center_fallback:
                    return None
                # Fallback crop when detector misses: centered portrait crop.
                img_h, img_w = img.shape[:2]
                crop_w = int(img_w * 0.64)
                crop_h = int(img_h * 0.82)
                nx1 = max(0, (img_w - crop_w) // 2)
                ny1 = max(0, int(img_h * 0.04))
                nx2 = min(img_w, nx1 + crop_w)
                ny2 = min(img_h, ny1 + crop_h)
                crop = img[ny1:ny2, nx1:nx2]
                if crop.size == 0:
                    return None
                if not self._is_head_shoulder_crop(crop):
                    return None
                snapshots_dir = os.path.join(self.visitor_reports_dir, 'snapshots')
                os.makedirs(snapshots_dir, exist_ok=True)
                snapshot_path = os.path.join(
                    snapshots_dir,
                    f"{visitor_code}_face_shoulder_{datetime.now().strftime('%Y%m%d%H%M%S')}.jpg",
                )
                cv2.imwrite(snapshot_path, crop)
                if self._image_has_valid_head_shoulder(snapshot_path):
                    return snapshot_path
                return None

            x, y, w, h = face_bbox
            img_h, img_w = img.shape[:2]
            # Keep crop focused on person portrait (face to shoulder), avoid over-expanding to torso.
            expand_x = int(w * 0.22)
            expand_y_top = int(h * 0.24)
            expand_y_bottom = int(h * 0.95)

            nx1 = max(0, x - expand_x)
            ny1 = max(0, y - expand_y_top)
            nx2 = min(img_w, x + w + expand_x)
            ny2 = min(img_h, y + h + expand_y_bottom)
            crop = img[ny1:ny2, nx1:nx2]
            if crop.size == 0:
                return None
            if not self._is_head_shoulder_crop(crop):
                # Mild second try with a bit more shoulder room.
                expand_x = int(w * 0.28)
                expand_y_top = int(h * 0.20)
                expand_y_bottom = int(h * 1.08)
                nx1 = max(0, x - expand_x)
                ny1 = max(0, y - expand_y_top)
                nx2 = min(img_w, x + w + expand_x)
                ny2 = min(img_h, y + h + expand_y_bottom)
                crop = img[ny1:ny2, nx1:nx2]
                if crop.size == 0 or not self._is_head_shoulder_crop(crop):
                    return None

            snapshots_dir = os.path.join(self.visitor_reports_dir, 'snapshots')
            os.makedirs(snapshots_dir, exist_ok=True)
            snapshot_path = os.path.join(
                snapshots_dir,
                f"{visitor_code}_face_shoulder_{datetime.now().strftime('%Y%m%d%H%M%S')}.jpg",
            )
            cv2.imwrite(snapshot_path, crop)
            if self._image_has_valid_head_shoulder(snapshot_path):
                return snapshot_path
            return None
        except Exception:
            return None

    def _image_has_valid_head_shoulder(self, image_path):
        if not image_path or not os.path.exists(image_path):
            return False
        img = cv2.imread(image_path)
        if img is None:
            return False
        return self._is_head_shoulder_crop(img)

    def _image_has_face(self, image_path):
        if not image_path or not os.path.exists(image_path):
            return False
        img = cv2.imread(image_path)
        if img is None:
            return False
        return self._detect_primary_face_bbox_with_retries(img) is not None

    def _latest_existing_snapshot(self, visitor_code, require_face=False):
        snapshots_dir = os.path.join(self.visitor_reports_dir, 'snapshots')
        if not os.path.isdir(snapshots_dir):
            return None
        patterns = [
            os.path.join(snapshots_dir, f"{visitor_code}_face_shoulder_*.jpg"),
            os.path.join(snapshots_dir, f"{visitor_code}_face_shoulder.jpg"),
        ]
        candidates = []
        for pattern in patterns:
            candidates.extend(glob(pattern))
        candidates = [path for path in candidates if os.path.exists(path)]
        if not candidates:
            return None
        candidates.sort(key=lambda path: os.path.getmtime(path), reverse=True)
        if not require_face:
            return candidates[0]
        for path in candidates:
            if self._image_has_valid_head_shoulder(path):
                return path
        return None

    def _select_best_snapshot(self, visitor, visitor_code, prefer_first=False, reference_time=None):
        candidates = self._resolve_visitor_image_candidates(
            visitor,
            prefer_first=prefer_first,
            reference_time=reference_time,
        )

        # Strict pass: only accept face-detected portrait crop.
        for image_path in candidates:
            snapshot_path = self._prepare_face_to_shoulder_snapshot(
                image_path,
                visitor_code,
                allow_center_fallback=False,
            )
            if snapshot_path and os.path.exists(snapshot_path):
                return snapshot_path

        # Reuse previously generated valid face snapshots.
        existing = self._latest_existing_snapshot(visitor_code, require_face=True)
        if existing and os.path.exists(existing):
            return existing

        # Last resort (still produce an image if nothing else is available).
        if candidates:
            fallback = self._prepare_face_to_shoulder_snapshot(
                candidates[0],
                visitor_code,
                allow_center_fallback=False,
            )
            if fallback and os.path.exists(fallback):
                return fallback
        return None

    def _build_summary_with_reportlab(self, filepath, title_text, subtitle_text, rows):
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        elements.append(Paragraph(title_text, styles['Title']))
        elements.append(Paragraph(subtitle_text, styles['Normal']))
        elements.append(Spacer(1, 12))

        table = Table(rows)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.75, colors.black),
        ]))
        elements.append(table)
        doc.build(elements)

    def _build_summary_with_fpdf(self, filepath, title_text, subtitle_text, rows):
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=12)
        pdf.add_page()
        pdf.set_font('Arial', 'B', 16)
        pdf.cell(0, 10, self._safe_text(title_text), ln=1)
        pdf.set_font('Arial', '', 11)
        pdf.cell(0, 8, self._safe_text(subtitle_text), ln=1)
        pdf.ln(3)

        col_count = max(1, len(rows[0]) if rows else 1)
        usable_width = 190
        col_width = usable_width / col_count

        for row_idx, row in enumerate(rows):
            if row_idx == 0:
                pdf.set_font('Arial', 'B', 10)
            else:
                pdf.set_font('Arial', '', 10)
            for cell in row:
                pdf.cell(col_width, 8, self._safe_text(cell), border=1, ln=0)
            pdf.ln(8)

        pdf.output(filepath)

    def _build_event_visitors_with_reportlab(self, filepath, title_text, subtitle_text, visitors_payload, management_payload):
        doc = SimpleDocTemplate(filepath, pagesize=A4, leftMargin=16, rightMargin=16, topMargin=18, bottomMargin=18)
        styles = getSampleStyleSheet()
        card_text_style = styles['Normal'].clone('CardText')
        card_text_style.fontName = 'Helvetica'
        card_text_style.fontSize = 7
        card_text_style.leading = 8
        elements = []

        elements.append(Paragraph(title_text, styles['Title']))
        elements.append(Paragraph(subtitle_text, styles['Normal']))
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("Visitor Summary", styles['Heading2']))
        elements.append(Spacer(1, 6))

        def _card_for_visitor(item):
            body = []
            body.append(Paragraph(f"<b>{item['visitor_id']}</b>", card_text_style))
            body.append(Spacer(1, 2))

            image_path = item.get('snapshot_path')
            if image_path and os.path.exists(image_path):
                body.append(Image(image_path, width=62, height=78))
            else:
                body.append(Paragraph("Snapshot unavailable", styles['Italic']))
            body.append(Spacer(1, 4))
            body.append(Paragraph(f"In: {item['first_in']}", card_text_style))
            body.append(Paragraph(f"Out: {item['last_out']}", card_text_style))
            body.append(Paragraph(f"Duration: {item['duration']}", card_text_style))

            card = Table([[body]], colWidths=[133])
            card.setStyle(TableStyle([
                ('BOX', (0, 0), (-1, -1), 0.7, colors.HexColor('#6b7280')),
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8fafc')),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            return card

        if not visitors_payload:
            elements.append(Paragraph("No visitors found in selected window.", styles['Normal']))
        else:
            page_size = 16
            for start in range(0, len(visitors_payload), page_size):
                if start > 0:
                    elements.append(PageBreak())
                    elements.append(Paragraph(title_text, styles['Title']))
                    elements.append(Paragraph(subtitle_text, styles['Normal']))
                    elements.append(Spacer(1, 10))
                    elements.append(Paragraph("Visitor Summary", styles['Heading2']))
                    elements.append(Spacer(1, 6))

                chunk = visitors_payload[start:start + page_size]
                row_data = []
                current_row = []
                for item in chunk:
                    current_row.append(_card_for_visitor(item))
                    if len(current_row) == 4:
                        row_data.append(current_row)
                        current_row = []
                if current_row:
                    while len(current_row) < 4:
                        current_row.append('')
                    row_data.append(current_row)

                grid = Table(row_data, colWidths=[136, 136, 136, 136], hAlign='LEFT')
                grid.setStyle(TableStyle([
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 0),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                    ('TOPPADDING', (0, 0), (-1, -1), 0),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]))
                elements.append(grid)

        elements.append(PageBreak())
        elements.append(Paragraph("Management", styles['Heading2']))
        if management_payload:
            mgmt_rows = [['Staff ID', 'Name', 'Department', 'Position', 'Status', 'Added']]
            for item in management_payload:
                mgmt_rows.append([
                    item.get('staff_id', ''),
                    item.get('name', ''),
                    item.get('department', '-'),
                    item.get('position', '-'),
                    item.get('status', ''),
                    item.get('created_at', ''),
                ])
            mgmt_table = Table(mgmt_rows, colWidths=[64, 96, 92, 92, 52, 110])
            mgmt_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#6b7280')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f8fafc'), colors.HexColor('#eef2f7')]),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(mgmt_table)
        else:
            elements.append(Paragraph("No staff data available for this event window.", styles['Normal']))

        doc.build(elements)

    def _build_event_visitors_with_fpdf(self, filepath, title_text, subtitle_text, visitors_payload, management_payload):
        pdf = FPDF()
        pdf.set_auto_page_break(auto=False)

        pdf.add_page()
        pdf.set_font('Arial', 'B', 16)
        pdf.cell(0, 10, self._safe_text(title_text), ln=1)

        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 7, 'Visitor Summary', ln=1)
        pdf.set_font('Arial', '', 8)
        pdf.multi_cell(0, 4, self._safe_text(subtitle_text))
        pdf.ln(1)

        def _compact_datetime(value):
            text = str(value or '').strip()
            if not text or text == '-':
                return '-'
            for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S'):
                try:
                    dt = datetime.strptime(text[:19], fmt)
                    return dt.strftime('%Y-%m-%d %H:%M')
                except Exception:
                    continue
            return text[:16]

        if not visitors_payload:
            pdf.set_font('Arial', '', 10)
            pdf.cell(0, 8, 'No visitors found in selected window.', ln=1)
        else:
            card_w = 45
            card_h = 58
            card_gap_x = 4
            card_gap_y = 3
            left_margin = 8
            top_offset = max(24, int(pdf.get_y()) + 2)

            for idx, visitor_item in enumerate(visitors_payload):
                slot = idx % 16
                if slot == 0 and idx > 0:
                    pdf.add_page()
                    pdf.set_font('Arial', 'B', 12)
                    pdf.cell(0, 7, 'Visitor Summary', ln=1)
                    pdf.set_font('Arial', '', 8)
                    pdf.multi_cell(0, 4, self._safe_text(subtitle_text))
                    pdf.ln(1)
                    top_offset = max(24, int(pdf.get_y()) + 2)

                row = slot // 4
                col = slot % 4
                card_x = left_margin + (col * (card_w + card_gap_x))
                card_y = top_offset + (row * (card_h + card_gap_y))

                pdf.set_draw_color(108, 117, 125)
                pdf.rect(card_x, card_y, card_w, card_h)

                pdf.set_xy(card_x + 2, card_y + 2)
                pdf.set_font('Arial', 'B', 8)
                pdf.cell(card_w - 4, 4, self._safe_text(visitor_item['visitor_id']), ln=1)

                image_path = visitor_item.get('snapshot_path')
                if image_path and os.path.exists(image_path):
                    try:
                        image_w = 24
                        image_h = 26
                        image_x = card_x + (card_w - image_w) / 2
                        image_y = card_y + 8
                        pdf.image(image_path, x=image_x, y=image_y, w=image_w, h=image_h)
                    except Exception:
                        pdf.set_xy(card_x + 2, card_y + 18)
                        pdf.set_font('Arial', 'I', 6)
                        pdf.cell(card_w - 4, 3, 'No snap', ln=1)
                else:
                    pdf.set_xy(card_x + 2, card_y + 18)
                    pdf.set_font('Arial', 'I', 6)
                    pdf.cell(card_w - 4, 3, 'No snap', ln=1)

                details = [
                    f"In: {_compact_datetime(visitor_item.get('first_in'))}",
                    f"Out: {_compact_datetime(visitor_item.get('last_out'))}",
                    f"Duration: {visitor_item['duration']}",
                ]
                pdf.set_font('Arial', '', 5.8)
                text_y = card_y + 35
                for line in details:
                    pdf.set_xy(card_x + 2, text_y)
                    pdf.multi_cell(card_w - 4, 3.0, self._safe_text(line))
                    text_y = pdf.get_y()

        pdf.add_page()
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 7, 'Management', ln=1)
        pdf.set_font('Arial', 'B', 8)
        headers = ['Staff ID', 'Name', 'Department', 'Position', 'Status']
        widths = [26, 48, 42, 42, 26]
        for idx, header in enumerate(headers):
            pdf.cell(widths[idx], 6, self._safe_text(header), border=1)
        pdf.ln(6)
        pdf.set_font('Arial', '', 8)
        if management_payload:
            for row in management_payload:
                values = [
                    row.get('staff_id', ''),
                    row.get('name', ''),
                    row.get('department', '-'),
                    row.get('position', '-'),
                    row.get('status', ''),
                ]
                for idx, value in enumerate(values):
                    pdf.cell(widths[idx], 6, self._safe_text(value), border=1)
                pdf.ln(6)
                if pdf.get_y() > 260:
                    pdf.add_page()
                    pdf.set_font('Arial', 'B', 8)
                    for idx, header in enumerate(headers):
                        pdf.cell(widths[idx], 6, self._safe_text(header), border=1)
                    pdf.ln(6)
                    pdf.set_font('Arial', '', 8)
        else:
            pdf.cell(0, 6, 'No staff data available for this event window.', ln=1)

        pdf.output(filepath)

    def _build_visitor_with_reportlab(
        self,
        filepath,
        visitor_code,
        first_in,
        last_out,
        duration_text,
        capture_date_text,
        visitor_image_path,
    ):
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph(f"Visitor ID: {visitor_code}", styles['Title']))
        elements.append(Spacer(1, 8))

        summary_table = Table([
            ['Date', capture_date_text],
            ['First In Time', first_in.strftime('%Y-%m-%d %H:%M:%S')],
            ['Last Out Time', last_out.strftime('%Y-%m-%d %H:%M:%S')],
            ['Total Duration', duration_text],
        ], colWidths=[150, 280])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.whitesmoke, colors.beige]),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 14))

        if visitor_image_path and os.path.exists(visitor_image_path):
            elements.append(Image(visitor_image_path, width=180, height=230))
        doc.build(elements)

    def _build_visitor_with_fpdf(
        self,
        filepath,
        visitor_code,
        first_in,
        last_out,
        duration_text,
        capture_date_text,
        visitor_image_path,
    ):
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=12)
        pdf.add_page()
        pdf.set_font('Arial', 'B', 16)
        pdf.cell(0, 10, self._safe_text(f"Visitor ID: {visitor_code}"), ln=1)

        pdf.set_font('Arial', 'B', 10)
        summary_rows = [
            ('Date', capture_date_text),
            ('First In Time', first_in.strftime('%Y-%m-%d %H:%M:%S')),
            ('Last Out Time', last_out.strftime('%Y-%m-%d %H:%M:%S')),
            ('Total Duration', duration_text),
        ]
        for key, value in summary_rows:
            pdf.cell(45, 8, self._safe_text(key), border=1)
            pdf.cell(130, 8, self._safe_text(value), border=1, ln=1)

        if visitor_image_path and os.path.exists(visitor_image_path):
            pdf.ln(8)
            try:
                pdf.image(visitor_image_path, w=62)
            except Exception:
                pass

        pdf.output(filepath)

    def _collect_event_visitors_payload(self, start_dt, end_dt):
        sessions = VisitorSession.query.filter(
            VisitorSession.entry_time <= end_dt,
            or_(VisitorSession.exit_time.is_(None), VisitorSession.exit_time >= start_dt),
        ).order_by(VisitorSession.entry_time.asc()).all()

        grouped = {}
        now_local = datetime.now()
        for session in sessions:
            visitor_db_id = session.visitor_id
            overlap_start = max(session.entry_time, start_dt)
            overlap_end = min(session.exit_time or now_local, end_dt)
            if overlap_end < overlap_start:
                continue
            item = grouped.setdefault(visitor_db_id, {
                'first_in': None,
                'last_out': None,
            })
            if item['first_in'] is None or overlap_start < item['first_in']:
                item['first_in'] = overlap_start
            if item['last_out'] is None or overlap_end > item['last_out']:
                item['last_out'] = overlap_end

        visitors_payload = []
        if grouped:
            visitors = Visitor.query.filter(Visitor.id.in_(grouped.keys())).all()
            visitors_by_id = {v.id: v for v in visitors}
            ordered_visitor_ids = sorted(
                grouped.keys(),
                key=lambda vid: grouped[vid]['first_in'] or datetime.max,
            )
            display_id_map = {
                visitor_db_id: f"ID{idx}"
                for idx, visitor_db_id in enumerate(ordered_visitor_ids, start=1)
            }

            for visitor_db_id in ordered_visitor_ids:
                summary = grouped[visitor_db_id]
                visitor = visitors_by_id.get(visitor_db_id)
                visitor_code = display_id_map.get(visitor_db_id, f'ID{visitor_db_id}')
                canonical_code = visitor.visitor_id if visitor else f'VISITOR-{visitor_db_id}'
                first_in = summary['first_in']
                last_out = summary['last_out']

                snapshot_path = None
                if visitor is not None:
                    snapshot_path = self._select_best_snapshot(
                        visitor,
                        canonical_code,
                        prefer_first=True,
                        reference_time=first_in,
                    )

                duration_seconds = 0
                if first_in and last_out:
                    duration_seconds = max(0, int((last_out - first_in).total_seconds()))

                visitors_payload.append({
                    'visitor_id': visitor_code,
                    'date': first_in.strftime('%Y-%m-%d') if first_in else '-',
                    'first_in': first_in.strftime('%Y-%m-%d %H:%M:%S') if first_in else '-',
                    'last_out': last_out.strftime('%Y-%m-%d %H:%M:%S') if last_out else '-',
                    'duration': self._format_duration(duration_seconds),
                    'snapshot_path': snapshot_path,
                })

        return visitors_payload

    @staticmethod
    def _build_report_title_subtitle(start_date, end_date, report_type='daily', event_name=None, event_id=None, visitor_count=0):
        title_text = f"Visitor Report ({str(report_type or 'daily').title()})"
        subtitle_text = f"Period: {start_date} to {end_date}"
        if event_name:
            subtitle_text = f"Event: {event_name} | {subtitle_text}"
        if event_id:
            subtitle_text = f"{subtitle_text} | Event ID: {event_id}"
        subtitle_text = f"{subtitle_text} | Visitors: {int(visitor_count or 0)}"
        return title_text, subtitle_text

    def _archive_event_outputs(
        self,
        source_path,
        event_name,
        event_id,
        start_date,
        end_date,
        visitors_payload,
        management_payload,
    ):
        if not event_name:
            return
        event_dir = self._event_artifacts_dir(event_name)
        archived_path = os.path.join(event_dir, os.path.basename(source_path))
        if os.path.abspath(archived_path) != os.path.abspath(source_path):
            shutil.copyfile(source_path, archived_path)
        self._save_event_data_csv(
            event_dir=event_dir,
            event_name=event_name,
            event_id=event_id,
            start_date=start_date,
            end_date=end_date,
            visitors_payload=visitors_payload,
            management_payload=management_payload,
        )

    def _build_event_visitors_with_excel(self, filepath, title_text, subtitle_text, visitors_payload, management_payload):
        self._ensure_excel_backend()

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Event Report'
        sheet.sheet_view.showGridLines = False

        for col_idx in range(1, 17):
            sheet.column_dimensions[get_column_letter(col_idx)].width = 15

        header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
        section_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
        card_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        table_header_fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
        thin = Side(border_style='thin', color='CBD5E1')
        card_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        sheet.merge_cells('A1:P1')
        sheet['A1'] = title_text
        sheet['A1'].font = Font(name='Calibri', size=18, bold=True, color='FFFFFF')
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet['A1'].fill = header_fill
        sheet.row_dimensions[1].height = 30

        sheet.merge_cells('A2:P2')
        sheet['A2'] = subtitle_text
        sheet['A2'].font = Font(name='Calibri', size=11, bold=False, color='FFFFFF')
        sheet['A2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        sheet['A2'].fill = header_fill
        sheet.row_dimensions[2].height = 36

        sheet.merge_cells('A4:P4')
        sheet['A4'] = 'Visitor Summary'
        sheet['A4'].font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
        sheet['A4'].alignment = Alignment(horizontal='left', vertical='center')
        sheet['A4'].fill = section_fill
        sheet.row_dimensions[4].height = 24

        base_row = 6
        card_rows = 16
        cards_per_row = 4

        if not visitors_payload:
            sheet.merge_cells(f'A{base_row}:P{base_row}')
            sheet[f'A{base_row}'] = 'No visitors found in selected window.'
            sheet[f'A{base_row}'].font = Font(name='Calibri', size=11, italic=True, color='334155')
        else:
            for idx, visitor_item in enumerate(visitors_payload):
                row_group = idx // cards_per_row
                col_group = idx % cards_per_row
                start_col = 1 + (col_group * 4)
                end_col = start_col + 3
                start_row = base_row + (row_group * card_rows)
                end_row = start_row + card_rows - 1

                for row_idx in range(start_row, end_row + 1):
                    if start_row + 2 <= row_idx <= start_row + 8:
                        sheet.row_dimensions[row_idx].height = 18
                    else:
                        sheet.row_dimensions[row_idx].height = max(sheet.row_dimensions[row_idx].height or 0, 16)
                    for col_idx in range(start_col, end_col + 1):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        cell.border = card_border
                        cell.fill = card_fill

                sheet.merge_cells(
                    f'{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{start_row}'
                )
                header_cell = sheet.cell(row=start_row, column=start_col)
                header_cell.value = visitor_item.get('visitor_id', 'ID')
                header_cell.font = Font(name='Calibri', size=11, bold=True, color='0F172A')
                header_cell.alignment = Alignment(horizontal='center', vertical='center')

                snapshot_path = visitor_item.get('snapshot_path')
                if snapshot_path and os.path.exists(snapshot_path):
                    try:
                        img = XLImage(snapshot_path)
                        img.width = 126
                        img.height = 132
                        anchor_col = min(16, start_col + 1)
                        img.anchor = f'{get_column_letter(anchor_col)}{start_row + 2}'
                        sheet.add_image(img)
                    except Exception:
                        sheet.merge_cells(
                            f'{get_column_letter(start_col)}{start_row + 5}:{get_column_letter(end_col)}{start_row + 5}'
                        )
                        miss_cell = sheet.cell(row=start_row + 5, column=start_col)
                        miss_cell.value = 'Snapshot unavailable'
                        miss_cell.font = Font(name='Calibri', size=9, italic=True, color='64748B')
                        miss_cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    sheet.merge_cells(
                        f'{get_column_letter(start_col)}{start_row + 5}:{get_column_letter(end_col)}{start_row + 5}'
                    )
                    miss_cell = sheet.cell(row=start_row + 5, column=start_col)
                    miss_cell.value = 'Snapshot unavailable'
                    miss_cell.font = Font(name='Calibri', size=9, italic=True, color='64748B')
                    miss_cell.alignment = Alignment(horizontal='center', vertical='center')

                details = [
                    f"In: {visitor_item.get('first_in', '-')}",
                    f"Out: {visitor_item.get('last_out', '-')}",
                    f"Duration: {visitor_item.get('duration', '-')}",
                ]
                for offset, text in enumerate(details):
                    detail_row = start_row + 11 + offset
                    sheet.merge_cells(
                        f'{get_column_letter(start_col)}{detail_row}:{get_column_letter(end_col)}{detail_row}'
                    )
                    detail_cell = sheet.cell(row=detail_row, column=start_col)
                    detail_cell.value = text
                    detail_cell.font = Font(name='Calibri', size=9, color='0F172A')
                    detail_cell.alignment = Alignment(horizontal='left', vertical='center')

        used_rows = 1 if not visitors_payload else ((len(visitors_payload) - 1) // cards_per_row + 1) * card_rows
        management_heading_row = base_row + used_rows + 2

        sheet.merge_cells(f'A{management_heading_row}:P{management_heading_row}')
        mgmt_heading_cell = sheet[f'A{management_heading_row}']
        mgmt_heading_cell.value = 'Management'
        mgmt_heading_cell.font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
        mgmt_heading_cell.alignment = Alignment(horizontal='left', vertical='center')
        mgmt_heading_cell.fill = section_fill
        sheet.row_dimensions[management_heading_row].height = 24

        table_start = management_heading_row + 2
        headers = ['Staff ID', 'Name', 'Department', 'Position', 'Status', 'Added']
        header_columns = [1, 4, 7, 10, 13, 15]
        header_spans = [3, 3, 3, 3, 2, 2]

        for idx, label in enumerate(headers):
            col = header_columns[idx]
            span = header_spans[idx]
            end_col = col + span - 1
            sheet.merge_cells(
                f'{get_column_letter(col)}{table_start}:{get_column_letter(end_col)}{table_start}'
            )
            cell = sheet.cell(row=table_start, column=col)
            cell.value = label
            cell.font = Font(name='Calibri', size=10, bold=True, color='1E293B')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = table_header_fill
            for c in range(col, end_col + 1):
                sheet.cell(row=table_start, column=c).border = card_border

        table_row = table_start + 1
        if management_payload:
            for item in management_payload:
                values = [
                    item.get('staff_id', ''),
                    item.get('name', ''),
                    item.get('department', '-'),
                    item.get('position', '-'),
                    item.get('status', ''),
                    item.get('created_at', ''),
                ]
                for idx, value in enumerate(values):
                    col = header_columns[idx]
                    span = header_spans[idx]
                    end_col = col + span - 1
                    sheet.merge_cells(
                        f'{get_column_letter(col)}{table_row}:{get_column_letter(end_col)}{table_row}'
                    )
                    cell = sheet.cell(row=table_row, column=col)
                    cell.value = value
                    cell.font = Font(name='Calibri', size=10, color='0F172A')
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    for c in range(col, end_col + 1):
                        sheet.cell(row=table_row, column=c).border = card_border
                table_row += 1
        else:
            sheet.merge_cells(f'A{table_row}:P{table_row}')
            cell = sheet[f'A{table_row}']
            cell.value = 'No staff data available for this event window.'
            cell.font = Font(name='Calibri', size=10, italic=True, color='64748B')
            cell.alignment = Alignment(horizontal='left', vertical='center')

        sheet.freeze_panes = 'A6'
        workbook.save(filepath)

    def generate_pdf_report(self, start_date, end_date, report_type='daily', event_name=None, event_id=None):
        self._ensure_pdf_backend()
        safe_event_name = (event_name or '').strip().replace(' ', '_')
        if safe_event_name:
            filename = f"Visitor_Report_{safe_event_name}_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        else:
            filename = f"Visitor_Report_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join(self.reports_dir, filename)

        s_date = self._parse_datetime(start_date, is_end=False)
        e_date = self._parse_datetime(end_date, is_end=True)
        if e_date < s_date:
            raise ValueError('end_date must be on/after start_date')

        visitors_payload = self._collect_event_visitors_payload(s_date, e_date)
        management_payload = self._collect_management_payload(e_date)
        title_text, subtitle_text = self._build_report_title_subtitle(
            start_date=start_date,
            end_date=end_date,
            report_type=report_type,
            event_name=event_name,
            event_id=event_id,
            visitor_count=len(visitors_payload),
        )

        if REPORTLAB_AVAILABLE:
            self._build_event_visitors_with_reportlab(
                filepath,
                title_text,
                subtitle_text,
                visitors_payload,
                management_payload,
            )
        else:
            self._build_event_visitors_with_fpdf(
                filepath,
                title_text,
                subtitle_text,
                visitors_payload,
                management_payload,
            )

        try:
            self._archive_event_outputs(
                source_path=filepath,
                event_name=event_name,
                event_id=event_id,
                start_date=start_date,
                end_date=end_date,
                visitors_payload=visitors_payload,
                management_payload=management_payload,
            )
        except Exception:
            # Report download should still succeed even if local archival fails.
            pass

        return filepath

    def generate_excel_report(self, start_date, end_date, report_type='daily', event_name=None, event_id=None):
        self._ensure_excel_backend()
        safe_event_name = (event_name or '').strip().replace(' ', '_')
        if safe_event_name:
            filename = f"Visitor_Report_{safe_event_name}_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        else:
            filename = f"Visitor_Report_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.reports_dir, filename)

        s_date = self._parse_datetime(start_date, is_end=False)
        e_date = self._parse_datetime(end_date, is_end=True)
        if e_date < s_date:
            raise ValueError('end_date must be on/after start_date')

        visitors_payload = self._collect_event_visitors_payload(s_date, e_date)
        management_payload = self._collect_management_payload(e_date)
        title_text, subtitle_text = self._build_report_title_subtitle(
            start_date=start_date,
            end_date=end_date,
            report_type=report_type,
            event_name=event_name,
            event_id=event_id,
            visitor_count=len(visitors_payload),
        )

        self._build_event_visitors_with_excel(
            filepath,
            title_text,
            subtitle_text,
            visitors_payload,
            management_payload,
        )

        try:
            self._archive_event_outputs(
                source_path=filepath,
                event_name=event_name,
                event_id=event_id,
                start_date=start_date,
                end_date=end_date,
                visitors_payload=visitors_payload,
                management_payload=management_payload,
            )
        except Exception:
            pass

        return filepath

    def generate_visitor_pdf(self, visitor, event_start=None, event_end=None):
        self._ensure_pdf_backend()
        if visitor is None:
            raise ValueError('visitor is required')

        sessions = VisitorSession.query.filter_by(visitor_id=visitor.id).order_by(VisitorSession.entry_time.asc()).all()
        if not sessions:
            raise ValueError('No sessions found for visitor')

        now_local = datetime.now()
        normalized_sessions = []
        for session in sessions:
            in_time = session.entry_time
            out_time = session.exit_time or visitor.last_seen or now_local
            if event_start:
                in_time = max(in_time, event_start)
            if event_end:
                out_time = min(out_time, event_end)
            if out_time < in_time:
                continue
            normalized_sessions.append((in_time, out_time, bool(session.exit_time)))

        if not normalized_sessions:
            raise ValueError('No session data found in selected event window')

        first_in = min(item[0] for item in normalized_sessions)
        last_out = max(item[1] for item in normalized_sessions)
        duration_seconds = max(0, int((last_out - first_in).total_seconds()))
        duration_text = self._format_duration(duration_seconds)
        capture_date_text = first_in.strftime('%Y-%m-%d')
        visitor_image_path = self._select_best_snapshot(
            visitor,
            visitor.visitor_id,
            prefer_first=True,
            reference_time=first_in,
        )

        filename = f"{visitor.visitor_id}_report.pdf"
        filepath = os.path.join(self.visitor_reports_dir, filename)

        if REPORTLAB_AVAILABLE:
            self._build_visitor_with_reportlab(
                filepath,
                visitor.visitor_id,
                first_in,
                last_out,
                duration_text,
                capture_date_text,
                visitor_image_path,
            )
        else:
            self._build_visitor_with_fpdf(
                filepath,
                visitor.visitor_id,
                first_in,
                last_out,
                duration_text,
                capture_date_text,
                visitor_image_path,
            )
        return filepath
